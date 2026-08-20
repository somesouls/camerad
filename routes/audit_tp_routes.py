# -*- coding: utf-8 -*-
"""audit_tp_routes.py — Menu "Audit Training Phrase".

Fitur berdiri sendiri (BUKAN bagian rail Step 1-16 Analisis Dialogflow).
Tujuan: higiene data intent. Menyandingkan training phrase ANTAR intent memakai
SBERT (model sama seperti pipeline) untuk menemukan:
  1) Konflik antar-intent  : frasa mirip yang berada di intent BERBEDA
                             (penyebab utama match rate turun / tumpang tindih).
  2) Duplikat dalam intent : frasa nyaris kembar di intent SAMA (seed mubazir).

Input = ZIP "Database Intent Dialogflow" (output Step 3/13) ATAU langsung file
xlsx Training Phrase berkolom ["ID", "Training Phrase"], ATAU "Gunakan Data
Terakhir" yang membaca katalog dari slot run tetap (auto-daily) yang diperbarui
otomatis tiap hari + tombol "Tarik ulang katalog".

Daftarkan dengan:  import audit_tp_routes; audit_tp_routes.register(app)

RBAC: path /audit-tp & /api/audit-tp/* jatuh ke area "dialogflow" (default),
aksi "read" — sama seperti tool Dialogflow lain, jadi tak perlu ubah izin.
"""
import io
import os
import base64

from fastapi import Request
from fastapi.responses import JSONResponse
from starlette.concurrency import run_in_threadpool

from app_core import render_page

try:
    import numpy as np
except Exception:
    np = None

import openpyxl


def find_latest_step3_zip():
    """Cari ZIP "Database Intent Dialogflow" terakhir dari run pipeline.

    Output Step 3 (avaya1) maupun Step 13 (avaya2) sama-sama menyimpan ZIP
    training+intent. pipeline_routes.save_artifact() menyimpan dict step dengan
    kunci "status": "done" (BUKAN "Selesai"; label "Selesai" hanya ada di dalam
    sub-dict summary). Karena itu deteksi memakai status == "done" dan
    memverifikasi file benar-benar ada di disk. Bila ada beberapa run, ambil
    yang "at"-nya paling baru; Step 3 dan Step 13 sama-sama diperiksa.
    """
    from app_core import CONFIG
    import pipeline_routes
    runs_dir = CONFIG.get("runs_dir", "")
    if not runs_dir or not os.path.isdir(runs_dir):
        return None
    latest_time = ""
    latest_zip = None
    try:
        for run_id in os.listdir(runs_dir):
            try:
                state = pipeline_routes.load_state(CONFIG, run_id)
            except Exception:
                continue
            if not state:
                continue
            steps = state.get("steps", {}) or {}
            for key in ("3", "13"):
                step = steps.get(key)
                if not step:
                    continue
                fname = step.get("file")
                if not fname:
                    continue
                if step.get("status") != "done":
                    continue
                if str(step.get("ext", "")).lower() not in ("zip", ""):
                    continue
                cand = os.path.join(runs_dir, run_id, fname)
                if not os.path.isfile(cand):
                    continue
                at = str(step.get("at", ""))
                # Tie-break: run_id ikut dibandingkan agar deterministik.
                marker = (at, run_id, key)
                if marker > (latest_time, "", "") or latest_zip is None:
                    if at >= latest_time:
                        latest_time = at
                        latest_zip = cand
    except Exception:
        pass
    return latest_zip

# ---------------------------------------------------------------------------
# Ambang default (bisa diubah dari form)
# ---------------------------------------------------------------------------
DEF_MIN_CROSS = float(os.environ.get("AUDIT_TP_MIN_CROSS", "0.85"))
DEF_MIN_DUP = float(os.environ.get("AUDIT_TP_MIN_DUP", "0.97"))
MAX_PAIRS = int(os.environ.get("AUDIT_TP_MAX_PAIRS", "2000"))
MAX_PHRASES = int(os.environ.get("AUDIT_TP_MAX_PHRASES", "8000"))
CHUNK = 256


def _default_encode(texts):
    """Pakai encoder SBERT yang sudah ada (knowledge_semantic). Return NxD
    ternormalisasi (float32) atau None bila model tak tersedia."""
    import knowledge_semantic as ksem
    if not ksem.is_available():
        return None
    return ksem._encode(list(texts))


def _read_training_rows(xlsx_bytes):
    """Baca xlsx Training Phrase -> list[(intent, phrase)]. Cari kolom
    'ID'/'Intent' & 'Training Phrase' dari baris header."""
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    header = None
    for r in rows_iter:
        if r and any(str(c or "").strip() for c in r):
            header = [str(c or "").strip() for c in r]
            break
    if not header:
        raise Exception("Sheet Training Phrase kosong.")
    low = [h.lower() for h in header]

    def _find(cands):
        for i, h in enumerate(low):
            if h in cands:
                return i
        return -1

    ci = _find({"id", "intent", "intent name", "nama intent"})
    cp = _find({"training phrase", "training_phrase", "phrase", "frasa", "training phrases"})
    if ci < 0 or cp < 0:
        raise Exception("Header wajib memuat kolom 'ID' (intent) dan 'Training Phrase'. "
                        "Ditemukan: " + ", ".join(header))
    def _cell(row, idx):
        v = row[idx] if idx < len(row) else None
        return "" if v is None else str(v).strip()

    out = []
    for r in rows_iter:
        if not r:
            continue
        intent = _cell(r, ci)
        phrase = _cell(r, cp)
        if intent and phrase:
            out.append((intent, phrase))
    wb.close()
    return out


def _category_of(intent):
    """Tentukan kategori sebuah intent untuk pemecahan batch.

    Nama intent Dialogflow lazim berpola hierarki (mis. "PPh - Tarif",
    "PPN.Faktur", "layanan/registrasi"). Ambil segmen pertama sebagai kategori.
    Bila tak ada pemisah, pakai nama intent apa adanya.
    """
    s = (intent or "").strip()
    if not s:
        return "(tanpa kategori)"
    for sep in (" - ", " / ", "/", "|", ".", "_"):
        if sep in s:
            head = s.split(sep, 1)[0].strip()
            if head:
                return head
    return s


def _batch_by_category(uniq, max_phrases):
    """Pecah daftar (intent, phrase) menjadi beberapa batch <= max_phrases.

    Frasa dari satu intent SELALU utuh dalam satu batch (agar deteksi duplikat
    dalam-intent tidak terpotong). Intent diurutkan per kategori supaya intent
    sekategori cenderung berada di batch yang sama (deteksi konflik antar-intent
    tetap berjalan di dalam batch). Batch diisi lintas-kategori sampai mendekati
    max_phrases agar tidak boros batch untuk kategori kecil.
    """
    by_intent = {}
    order = []
    for intent, phrase in uniq:
        if intent not in by_intent:
            by_intent[intent] = []
            order.append(intent)
        by_intent[intent].append((intent, phrase))
    order.sort(key=lambda it: (_category_of(it), it))
    batches = []
    cur = []
    for intent in order:
        rows_i = by_intent[intent]
        if cur and (len(cur) + len(rows_i) > max_phrases):
            batches.append(cur)
            cur = []
        cur.extend(rows_i)
        if len(cur) >= max_phrases:
            batches.append(cur)
            cur = []
    if cur:
        batches.append(cur)
    return batches


def _score_pairs(uniq, min_cross, min_dup, chunk, encode_fn):
    """Pemindaian pasangan inti atas satu batch (intent, phrase) yang sudah unik.

    Semua embedding diasumsikan ternormalisasi sehingga cosine = dot product.
    Return (cross_raw, dup_raw):
      cross_raw = list[(score, intent_a, phrase_a, intent_b, phrase_b)]
      dup_raw   = list[(score, intent, phrase_a, phrase_b)]
    """
    n = len(uniq)
    if n < 2:
        return [], []
    intents = [x[0] for x in uniq]
    phrases = [x[1] for x in uniq]
    enc = encode_fn or _default_encode
    emb = enc(phrases)
    if emb is None:
        raise Exception("Model SBERT tidak aktif di server (torch / sentence-transformers). "
                        "Aktifkan KNOWLEDGE_SEMANTIC & pasang dependensi GPU/CPU.")
    emb = np.asarray(emb, dtype="float32")

    cross_seen = set()
    cross_raw = []
    dup_raw = []
    intent_arr = np.array(intents, dtype=object)

    for a in range(0, n, chunk):
        b = min(a + chunk, n)
        sims = emb[a:b] @ emb.T                     # (chunk x n)
        for local_i in range(b - a):
            i = a + local_i
            srow = sims[local_i]
            srow[i] = -1.0                          # buang self
            same = (intent_arr == intents[i])
            # --- konflik antar-intent (intent berbeda, >= min_cross) ---
            cross_mask = (~same) & (srow >= min_cross)
            for j in np.nonzero(cross_mask)[0]:
                j = int(j)
                key = (i, j) if i < j else (j, i)
                if key in cross_seen:
                    continue
                cross_seen.add(key)
                cross_raw.append((float(srow[j]), intents[i], phrases[i],
                                  intents[j], phrases[j]))
            # --- duplikat dalam intent (intent sama, >= min_dup) ---
            dup_mask = same & (srow >= min_dup)
            for j in np.nonzero(dup_mask)[0]:
                j = int(j)
                if j <= i:
                    continue
                dup_raw.append((float(srow[j]), intents[i], phrases[i], phrases[j]))
    return cross_raw, dup_raw


def audit_phrases(rows, min_cross=DEF_MIN_CROSS, min_dup=DEF_MIN_DUP,
                  top=MAX_PAIRS, chunk=CHUNK, encode_fn=None):
    """Inti algoritma. rows = list[(intent, phrase)]. encode_fn injectable untuk
    pengujian (default: SBERT).

    Bila jumlah frasa unik melebihi MAX_PHRASES, data dipecah per kategori
    menjadi beberapa batch (lihat _batch_by_category) lalu hasilnya digabung —
    menggantikan error lama "Terlalu banyak training phrase". Konsekuensinya,
    konflik yang kebetulan jatuh di batch berbeda tidak dibandingkan; ini
    trade-off yang disengaja agar audit tetap muat di memori/CPU.
    """
    if np is None:
        raise Exception("numpy tidak tersedia di server.")
    # Dedup pasangan (intent, phrase) identik.
    seen = set()
    uniq = []
    for intent, phrase in rows:
        key = (intent, phrase)
        if key in seen:
            continue
        seen.add(key)
        uniq.append((intent, phrase))
    n = len(uniq)
    intents_all = [x[0] for x in uniq]
    if n < 2:
        return {"ok": True, "n_intents": len(set(intents_all)), "n_phrases": n,
                "min_cross": min_cross, "min_dup": min_dup,
                "cross_pairs": [], "intent_conflicts": [], "dupes": [],
                "note": "Terlalu sedikit frasa untuk diaudit."}

    # Pecah per kategori bila melebihi ambang; selain itu proses sekaligus.
    if n > MAX_PHRASES:
        batches = _batch_by_category(uniq, MAX_PHRASES)
    else:
        batches = [uniq]
    batched = len(batches) > 1

    cross_raw = []
    dup_raw = []
    for batch in batches:
        c, d = _score_pairs(batch, min_cross, min_dup, chunk, encode_fn)
        cross_raw.extend(c)
        dup_raw.extend(d)

    cross_raw.sort(key=lambda t: t[0], reverse=True)
    dup_raw.sort(key=lambda t: t[0], reverse=True)
    truncated = len(cross_raw) > top

    # Agregasi per pasangan intent (dari seluruh konflik sebelum truncation).
    agg = {}
    for s, ia, pa, ib, pb in cross_raw:
        pk = (ia, ib) if ia <= ib else (ib, ia)
        cur = agg.get(pk)
        if cur is None:
            agg[pk] = [1, s]
        else:
            cur[0] += 1
            if s > cur[1]:
                cur[1] = s
    intent_conflicts = [{"intent_a": k[0], "intent_b": k[1],
                         "count": v[0], "max_score": round(v[1], 4)}
                        for k, v in agg.items()]
    intent_conflicts.sort(key=lambda d: (d["count"], d["max_score"]), reverse=True)

    cross_out = [{"intent_a": ia, "phrase_a": pa, "intent_b": ib,
                  "phrase_b": pb, "score": round(s, 4)}
                 for s, ia, pa, ib, pb in cross_raw[:top]]
    dup_out = [{"intent": ia, "phrase_a": pa, "phrase_b": pb,
                "score": round(s, 4)} for s, ia, pa, pb in dup_raw[:top]]

    result = {"ok": True, "n_intents": len(set(intents_all)), "n_phrases": n,
              "min_cross": min_cross, "min_dup": min_dup, "truncated": truncated,
              "cross_pairs": cross_out, "intent_conflicts": intent_conflicts,
              "dupes": dup_out}
    if batched:
        result["batched"] = True
        result["n_batches"] = len(batches)
        result["note"] = ("Data besar (%d frasa, %d intent) dipecah per kategori "
                          "menjadi %d batch; konflik lintas-batch tidak dibandingkan."
                          % (n, len(set(intents_all)), len(batches)))
    return result


def build_report_xlsx(result):
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Konflik Antar-Intent"
    ws1.append(["Skor", "Intent A", "Frasa A", "Intent B", "Frasa B"])
    for p in result.get("cross_pairs", []):
        ws1.append([p["score"], p["intent_a"], p["phrase_a"], p["intent_b"], p["phrase_b"]])
    ws2 = wb.create_sheet("Ringkasan Pasangan Intent")
    ws2.append(["Intent A", "Intent B", "Jumlah Konflik", "Skor Maks"])
    for c in result.get("intent_conflicts", []):
        ws2.append([c["intent_a"], c["intent_b"], c["count"], c["max_score"]])
    ws3 = wb.create_sheet("Duplikat Dalam Intent")
    ws3.append(["Skor", "Intent", "Frasa A", "Frasa B"])
    for d in result.get("dupes", []):
        ws3.append([d["score"], d["intent"], d["phrase_a"], d["phrase_b"]])
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def run_audit(file_bytes, min_cross=DEF_MIN_CROSS, min_dup=DEF_MIN_DUP):
    """Terima bytes upload (ZIP Step 3/13 atau xlsx Training Phrase), audit,
    dan lampirkan xlsx laporan (base64)."""
    train_bytes = None
    try:
        import pipeline_routes
        train_bytes, _content = pipeline_routes.extract_training_intent(file_bytes)
    except Exception:
        train_bytes = file_bytes           # fallback: dianggap xlsx langsung
    rows = _read_training_rows(train_bytes)
    if not rows:
        raise Exception("Tidak ada baris training phrase yang terbaca.")
    result = audit_phrases(rows, min_cross=min_cross, min_dup=min_dup)
    try:
        xlsx = build_report_xlsx(result)
        result["xlsx_b64"] = base64.b64encode(xlsx).decode("ascii")
        result["xlsx_name"] = "audit_training_phrase.xlsx"
    except Exception:
        pass
    return result


# ---------------------------------------------------------------------------
# Refresh katalog headless (Step 3) — dipakai penjadwal harian & tombol manual
# ---------------------------------------------------------------------------
_SCHED_STARTED = False


def run_step3_headless(cfg, run_id):
    """Tarik katalog intent + training phrase (Step 3) TANPA request HTTP.

    Memanggil ulang fungsi Step 3 dari pipeline (pipeline_routes) dengan Ctx
    minimal, memakai service account (access_token kosong), lalu menyimpan ZIP
    ke slot run tetap. Menu Audit membacanya lewat "Gunakan Data Terakhir".

    CATATAN PENTING: log harian (menu Kelola Data Dialogflow) HANYA memuat
    percakapan/logs — BUKAN training phrase. Jadi katalog intent+training phrase
    WAJIB ditarik terpisah langsung dari Dialogflow Agent API di sini.
    """
    import pipeline_routes
    os.makedirs(pipeline_routes.run_dir(cfg, run_id), exist_ok=True)
    ctx = pipeline_routes.Ctx(run_id, {}, {}, {})
    pipeline_routes.step3_training_intent(cfg, ctx)
    summary, at = {}, ""
    try:
        st = pipeline_routes.load_state(cfg, run_id) or {}
        step = (st.get("steps", {}) or {}).get("3", {}) or {}
        summary = step.get("summary", {}) or {}
        at = step.get("at", "")
    except Exception:
        pass
    return {"at": at, "summary": summary, "run": run_id}


def _start_catalog_scheduler():
    """Nyalakan penjadwal harian refresh katalog. Dipasang dari register() agar
    tidak perlu menyentuh web_app.py (mengikuti pola modul lain di proyek ini,
    mis. awe_analytics dipasang dari studio_routes). Idempoten: aman meski
    register() terpanggil lebih dari sekali."""
    global _SCHED_STARTED
    if _SCHED_STARTED:
        return
    if (os.environ.get("PIPELINE_SCHEDULER", "1") or "1").strip() == "0":
        return
    if (os.environ.get("PIPELINE_CATALOG_REFRESH", "1") or "1").strip() == "0":
        print("[audit-tp] refresh katalog otomatis dimatikan (PIPELINE_CATALOG_REFRESH=0).", flush=True)
        return
    try:
        from apscheduler.schedulers.background import BackgroundScheduler
    except Exception as e:
        print("[audit-tp] APScheduler belum terpasang, refresh katalog otomatis dilewati:", e, flush=True)
        return
    try:
        from zoneinfo import ZoneInfo
        tz = ZoneInfo("Asia/Jakarta")
    except Exception:
        tz = None
    hour = int(os.environ.get("PIPELINE_CATALOG_HOUR",
                              os.environ.get("PIPELINE_INGEST_HOUR", "8")))
    minute = int(os.environ.get("PIPELINE_CATALOG_MINUTE",
                                os.environ.get("PIPELINE_INGEST_MINUTE", "0")))
    run_id = os.environ.get("PIPELINE_CATALOG_RUN", "auto-daily")

    def _job():
        from app_core import CONFIG
        try:
            run_step3_headless(CONFIG, run_id)
            print("[audit-tp] katalog Step 3 diperbarui -> run '%s'." % run_id, flush=True)
        except Exception as e:
            print("[audit-tp] refresh katalog gagal:", e, flush=True)

    try:
        sch = BackgroundScheduler(timezone=tz) if tz else BackgroundScheduler()
        sch.add_job(_job, "cron", hour=hour, minute=minute,
                    id="audit_tp_catalog_refresh", replace_existing=True)
        sch.start()
        globals()["_CATALOG_SCH"] = sch      # jaga referensi agar tak di-GC
        _SCHED_STARTED = True
        print("[audit-tp] refresh katalog otomatis aktif jam %02d:%02d Asia/Jakarta -> run '%s'."
              % (hour, minute, run_id), flush=True)
    except Exception as e:
        print("[audit-tp] gagal memulai penjadwal katalog:", e, flush=True)


# ---------------------------------------------------------------------------
# Route
# ---------------------------------------------------------------------------
async def audit_tp_page(request: Request):
    return render_page(request, "audit_tp.html", "audit_tp")


def _to_float(v, default):
    try:
        return float(v)
    except Exception:
        return default


async def api_audit_run(request: Request):
    form = await request.form()

    use_latest = (form.get("use_latest") == "1")
    data = None

    if use_latest:
        zip_path = find_latest_step3_zip()
        if not zip_path or not os.path.isfile(zip_path):
            return JSONResponse({"ok": False, "error": "Tidak ditemukan data Step 3/13 terakhir. Klik 'Tarik ulang katalog', unggah file, atau jalankan Step 3 di Analisis Dialogflow."})
        with open(zip_path, "rb") as f:
            data = f.read()
    else:
        up = form.get("file")
        if up is None or not hasattr(up, "read"):
            return JSONResponse({"ok": False, "error": "Unggah file ZIP/xlsx training phrase atau centang Gunakan Data Terakhir."})
        data = await up.read()

    min_cross = _to_float(form.get("min_cross"), DEF_MIN_CROSS)
    min_dup = _to_float(form.get("min_dup"), DEF_MIN_DUP)

    def _run():
        return run_audit(data, min_cross=min_cross, min_dup=min_dup)
    try:
        return JSONResponse(await run_in_threadpool(_run))
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)})


async def api_audit_refresh(request: Request):
    """Tarik ulang katalog Step 3 sekarang juga (tombol manual di menu Audit).
    Hanya penarikan katalog intent+training phrase — BUKAN menjalankan seluruh
    Analisis Dialogflow."""
    from app_core import CONFIG
    run_id = os.environ.get("PIPELINE_CATALOG_RUN", "auto-daily")

    def _run():
        info = run_step3_headless(CONFIG, run_id)
        return {"ok": True, **info}
    try:
        return JSONResponse(await run_in_threadpool(_run))
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)})


def register(app):
    app.add_api_route("/audit-tp", audit_tp_page, methods=["GET"])
    app.add_api_route("/api/audit-tp/run", api_audit_run, methods=["POST"])
    app.add_api_route("/api/audit-tp/refresh", api_audit_refresh, methods=["POST"])
    _start_catalog_scheduler()
