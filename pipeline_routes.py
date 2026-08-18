# -*- coding: utf-8 -*-
"""pipeline_routes.py — Lapisan rute UI/proxy pipeline Dialogflow + Avaya.
Konteks request (Ctx/build_ctx), dispatch aksi, unduhan artefak, chat LLM, dan
rute (/,/tools,/api/config,/api/chat). Implementasi langkah dipindah ke
pipeline_steps.py; helper leaf ke pipeline_helpers.py. Nama helper & fungsi
langkah di-re-export ke namespace modul ini agar pemanggilan
`pipeline_routes.<nama>` dari modul lain (mis. web_app.py -> awe_routes) tetap
bekerja seperti sebelum refactor.

MIGRASI PENYIMPANAN: "Run ID" manual diganti "dataset" (kunci = rentang tanggal
log + bahasa, mis. 2026-08-01__2026-08-15__id). Step 1 membuat/mengaktifkan
dataset dari form-nya; aksi lain memakai dataset AKTIF bila run kosong. Reset =
soft-archive (bisa dimuat ulang). Unduhan diregenerasi dari DB (sumber
kebenaran) sehingga editan Step 6/9 langsung ikut terunduh.

Daftarkan dengan:  import pipeline_routes; pipeline_routes.register(app)
"""
import os
import re
import io
import csv
import json
import time
import zipfile
import datetime as _dt
from typing import Optional

import requests
from fastapi import FastAPI, Request
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from fastapi.responses import HTMLResponse, JSONResponse, Response, PlainTextResponse, RedirectResponse
from starlette.datastructures import UploadFile as StarletteUploadFile
from starlette.concurrency import run_in_threadpool
from urllib.parse import quote as _quote

import llm_client
import users_db as usr
import analytics_db as adb
import avaya_db as avdb
import avaya_client as avc
import threading as _threading
import uuid as _uuid
import ingest
import glossary_db as gdb
import disambig_db as ddb
import intentmap_db as imdb
import knowledge_ctx as kctx  # gabungan konteks 3 pustaka utk prompt
import pustaka_stats as pstats  # statistik pemakaian pustaka
import intent_describe as idesc  # job deskripsi AI utk katalog intent
import pii_mask  # masking PII sebelum teks dikirim ke LLM (Fase A)
import pipeline_store as pstore  # penyimpanan persisten (DB) pengganti Run ID

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

try:
    from dotenv import load_dotenv
    load_dotenv()
except Exception:
    pass

# --- Fondasi bersama dipindah ke app_core.py (migrasi bertahap, langkah 1) ---

from app_core import CONFIG, XLSX_MIME, BASE_DIR, render_page, _load_html

# --- Helper leaf (state/artefak, Google auth, HTTP, XLSX) dipindah ke
#     pipeline_helpers.py. Diimpor ke namespace ini agar pemanggilan
#     `pipeline_routes.<helper>` dari modul lain tetap bekerja. ---
from pipeline_helpers import (
    run_dir, state_path, load_state, save_state, set_step, get_state,
    reset_run, save_ngrok, mime_for_ext, save_artifact, resolve_input_bytes,
    read_upload, google_token, http_post_json, http_get_json,
    resolve_api_base, api_endpoint, _api_headers, curl_multipart,
    curl_multipart_raw, curl_json_raw, curl_get_raw, curl_post_json,
    xlsx_build, zip_build, _wb_from_bytes, sheet_headers, read_sheet,
    _sv, _is_numeric, _find_header, xlsx_upsert_sheet,
    artifact_bytes, artifact_meta,
)

# --- Implementasi langkah (step1..step16, avaya*, helper langkah) dipindah ke
#     pipeline_steps.py. Re-export semua nama publik + helper underscore ke
#     namespace ini agar dispatch & pemanggilan luar tetap bekerja. ---
from pipeline_steps import *  # noqa: F401,F403
from pipeline_steps import (
    _s2_match, _intent_freq_map, _sync_catalog_safe, _csv_bytes,
    _avaya_summary_from_result, _avaya_persist, _avaya_inputs,
)


# =============================================================
# Konteks request (pengganti $_POST / $_GET / $_FILES)
# =============================================================
class Ctx:
    def __init__(self, run, form, files, query):
        self.run = run
        self.form = form      # dict[str,str]
        self.files = files    # dict[str, list[(bytes, filename)]]
        self.query = query    # dict[str,str]

    def P(self, name, default=""):
        v = self.form.get(name)
        return default if v is None else v

    def G(self, name, default=""):
        v = self.query.get(name)
        return default if v is None else v

    def R(self, name, default=""):
        if name in self.form:
            return self.form[name]
        if name in self.query:
            return self.query[name]
        return default

    def file(self, field):
        lst = self.files.get(field)
        if lst:
            return lst[0]  # (bytes, filename)
        return None

    def file_list(self, field):
        return self.files.get(field, [])


async def build_ctx(request: Request) -> Ctx:
    query = dict(request.query_params)
    form_fields = {}
    files = {}
    if request.method == "POST":
        form = await request.form()
        for key, value in form.multi_items():
            if isinstance(value, StarletteUploadFile):
                data = await value.read()
                files.setdefault(key, []).append((data, value.filename or ""))
            else:
                form_fields[key] = value  # last-wins, seperti $_POST
    # 'run' kini opsional (kunci dataset). Bila kosong, aksi memakai dataset
    # AKTIF (di-resolve saat dispatch). Bila diisi, harus format aman.
    run = query.get("run") or form_fields.get("run") or ""
    if run and not re.match(r"^[A-Za-z0-9_\-]{1,64}$", run):
        raise Exception("Run ID tidak valid.")
    return Ctx(run, form_fields, files, query)


# =============================================================
# Dataset helpers (pengganti "Run ID")
# =============================================================
def _pconn():
    return pstore.init_db(pstore.connect())


def _active_run(cfg):
    """dkey dataset aktif (paling baru disentuh) atau "" bila belum ada."""
    conn = _pconn()
    try:
        ds = pstore.get_active_dataset(conn)
        return (ds.get("dkey") if ds else "") or ""
    finally:
        conn.close()


def _dataset_exists(run):
    """True bila 'run' (dkey) cocok dengan dataset yang BENAR-BENAR ada di DB."""
    if not run:
        return False
    conn = _pconn()
    try:
        return pstore.get_dataset_by_key(conn, str(run)) is not None
    finally:
        conn.close()


def _resolve_run_or_active(cfg, run):
    """Kembalikan 'run' bila dataset-nya ada; kalau tidak, pakai dataset AKTIF.

    Melindungi dari nilai RUN basi di sisi browser (mis. Run ID lama pra-migrasi
    yang masih tersimpan di localStorage, atau file JS lama yang ter-cache)
    yang tidak lagi cocok dataset mana pun. Tanpa ini, unduhan Step 1 akan 404
    dan input step (mis. Step 2) melapor 'Hasil Step N belum tersedia' padahal
    datanya ada di DB di bawah dataset AKTIF.
    """
    if run and _dataset_exists(run):
        return run
    return _active_run(cfg)


def _prep_step1_run(cfg, ctx):
    """Step 1 menentukan dataset dari form (rentang tanggal + bahasa). Buat/
    aktifkan dataset lalu set ctx.run = dkey. Bila field belum lengkap, biarkan
    step1_pull_logs sendiri yang memvalidasi & melempar pesan aslinya."""
    start = (ctx.P("start_date", "") or "").strip()
    end = (ctx.P("end_date", "") or "").strip()
    lang = ((ctx.P("bahasa", "id") or "id").strip().lower()) or "id"
    if not start or not end:
        return
    conn = _pconn()
    try:
        ds = pstore.get_or_create_dataset(conn, start, end, lang)
    finally:
        conn.close()
    ctx.run = (ds.get("dkey") if ds else "") or pstore.make_key(start, end, lang)


def _datasets_result(cfg):
    conn = _pconn()
    try:
        rows = pstore.list_datasets(conn, include_archived=True)
        active = pstore.get_active_dataset(conn)
    finally:
        conn.close()
    active_key = (active.get("dkey") if active else "") or ""
    out = []
    for d in rows:
        out.append({
            "run": d.get("dkey", ""),
            "label": d.get("label", ""),
            "range_start": d.get("range_start", ""),
            "range_end": d.get("range_end", ""),
            "lang": d.get("lang", ""),
            "status": d.get("status", ""),
            "updated_at": d.get("updated_at", ""),
            "active": (d.get("dkey", "") == active_key),
        })
    return {"ok": True, "datasets": out, "active": active_key}


def _activate_result(cfg, ctx):
    run = ctx.run
    if not run:
        raise Exception("Dataset tidak ditentukan.")
    conn = _pconn()
    try:
        ds = pstore.get_dataset_by_key(conn, run)
        if not ds:
            raise Exception("Dataset tidak ditemukan: %s" % run)
        pstore.activate_dataset(conn, ds["id"])
    finally:
        conn.close()
    st = load_state(cfg, run)
    st["ok"] = True
    st["run"] = run
    return st


def _state_result(cfg, run):
    st = load_state(cfg, run)
    if not isinstance(st, dict):
        st = {"run": run, "steps": {}}
    st["ok"] = True
    return st


# =============================================================
# DISPATCH (router — sama dgn switch($action) di index.php)
# =============================================================
def dispatch(action, cfg, ctx):
    # --- Aksi manajemen dataset (tak butuh run valid dari frontend) ---
    if action == "datasets":
        return _datasets_result(cfg)
    if action == "activate":
        return _activate_result(cfg, ctx)
    # --- Step 1 menetapkan dataset dari form-nya ---
    if action == "step1":
        _prep_step1_run(cfg, ctx)
    else:
        # --- Aksi lain: bila run kosong ATAU tidak cocok dataset yang ada
        #     (nilai RUN basi dari browser/localStorage/JS ter-cache), jatuh ke
        #     dataset AKTIF supaya unduhan & input step tetap ketemu datanya. ---
        ctx.run = _resolve_run_or_active(cfg, ctx.run)

    handlers = {
        "state": lambda: _state_result(cfg, ctx.run),
        "reset": lambda: reset_run(cfg, ctx.run),
        "step1": lambda: step1_pull_logs(cfg, ctx),
        "step2": lambda: step2_json_to_xlsx(cfg, ctx),
        "step3": lambda: step3_training_intent(cfg, ctx),
        "step4": lambda: step4_analyze(cfg, ctx),
        "step5": lambda: step5_qwen_judge(cfg, ctx),
        "step6load": lambda: step6_load(cfg, ctx),
        "step6": lambda: step6_save(cfg, ctx),
        "step7": lambda: step7_mkta(cfg, ctx),
        "step8load": lambda: step8_load(cfg, ctx),
        "step8": lambda: step8_run(cfg, ctx),
        "step9load": lambda: step9_load(cfg, ctx),
        "step9": lambda: step9_save(cfg, ctx),
        "step10": lambda: step10_build(cfg, ctx),
        "step11": lambda: step11_update(cfg, ctx),
        "step12": lambda: avaya1_upload_json(cfg, ctx),
        "step13": lambda: avaya2_pull_intents(cfg, ctx),
        "step14": lambda: avaya3_analyze(cfg, ctx),
        "step14start": lambda: avaya3_start(cfg, ctx),
        "step14progress": lambda: avaya3_progress(cfg, ctx),
        "step14fetch": lambda: avaya3_fetch(cfg, ctx),
        "step15": lambda: avaya4_dashboard(cfg, ctx),
        "step16": lambda: avaya5_excel(cfg, ctx),
        "avayadiag": lambda: avaya_diag(cfg, ctx),
    }
    fn = handlers.get(action)
    if fn is None:
        raise Exception("Aksi tidak dikenal: %s" % action)
    return fn()


# =============================================================
# DOWNLOAD (regenerasi dari DB = sumber kebenaran)
# =============================================================
def _bytes_resp(data, mime, filename=None):
    headers = {}
    if filename:
        headers["Content-Disposition"] = 'attachment; filename="%s"' % filename.replace('"', "")
    return Response(content=data, media_type=mime, headers=headers)


def _sheet_to_csv_bytes(xlsx_bytes, sheet_name):
    """Ekstrak 1 sheet dari XLSX (bytes) menjadi CSV (bytes). None bila gagal."""
    try:
        wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    except Exception:
        return None
    if sheet_name not in wb.sheetnames:
        return None
    ws = wb[sheet_name]
    out = io.StringIO()
    w = csv.writer(out)
    for row in ws.iter_rows(values_only=True):
        w.writerow(["" if c is None else c for c in row])
    return out.getvalue().encode("utf-8-sig")


def handle_download(cfg, request: Request):
    q = request.query_params
    run = q.get("run", "")
    if run and not re.match(r"^[A-Za-z0-9_\-]{1,64}$", run):
        return PlainTextResponse("Run ID tidak valid.", status_code=400)
    if not run:
        run = _active_run(cfg)
    elif not _dataset_exists(run):
        # RUN basi (tak cocok dataset mana pun, mis. Run ID lama pra-migrasi
        # yang masih ada di localStorage) -> pakai dataset AKTIF.
        run = _active_run(cfg)
    part = q.get("part", "")
    d = run_dir(cfg, run, create=False)

    def file_resp(path, mime, filename=None):
        with open(path, "rb") as f:
            data = f.read()
        return _bytes_resp(data, mime, filename)

    # --- CSV Step 10 (LM / Pembaruan): cache dulu, kalau tak ada regen dari DB ---
    if part in ("lm", "pembaruan"):
        m = {"lm": ("step10_lm.csv", "LM.csv", "LM"),
             "pembaruan": ("step10_pembaruan.csv", "Pembaruan.csv", "Pembaruan")}
        cache_name, dl_name, sheet = m[part]
        p = os.path.join(d, cache_name)
        if os.path.isfile(p):
            return file_resp(p, "text/csv; charset=utf-8", dl_name)
        xb = artifact_bytes(cfg, run, 10)
        if xb:
            data = _sheet_to_csv_bytes(xb, sheet)
            if data is not None:
                return _bytes_resp(data, "text/csv; charset=utf-8", dl_name)
        return PlainTextResponse("File CSV belum tersedia. Jalankan Step 10.", status_code=404)

    # --- ZIP usersays Step 11 ---
    if part == "zip11":
        p = os.path.join(d, "step11_usersays.zip")
        if os.path.isfile(p):
            return file_resp(p, "application/zip", "usersays_updated.zip")
        meta = artifact_meta(cfg, run, 11)
        if meta and (meta.get("ext") == "zip"):
            b = artifact_bytes(cfg, run, 11)
            if b:
                return _bytes_resp(b, "application/zip", meta.get("name") or "usersays_updated.zip")
        return PlainTextResponse("ZIP belum dibuat. Jalankan Step 11.", status_code=404)

    # --- Dashboard Avaya Step 15 ---
    if part == "avayadash":
        p = os.path.join(d, "step15_dashboard.html")
        if os.path.isfile(p):
            with open(p, "rb") as f:
                return Response(content=f.read(), media_type="text/html; charset=utf-8")
        b = artifact_bytes(cfg, run, 15)
        if b:
            return Response(content=b, media_type="text/html; charset=utf-8")
        return PlainTextResponse("Dashboard belum dibuat. Jalankan Step 15.", status_code=404)

    # --- Artefak per step (Excel/JSON/ZIP): regenerasi dari DB ---
    try:
        n = int(q.get("step", "0"))
    except Exception:
        n = 0
    b = artifact_bytes(cfg, run, n)
    meta = artifact_meta(cfg, run, n)
    if b is not None and meta is not None:
        return _bytes_resp(b, meta.get("mime") or "application/octet-stream",
                           str(meta.get("name") or ""))
    # fallback: cache disk (kompat lama)
    state = load_state(cfg, run)
    step = state["steps"].get(str(n))
    if step and step.get("file"):
        p = os.path.join(d, step["file"])
        if os.path.isfile(p):
            return file_resp(p, step.get("mime", "application/octet-stream"),
                             str(step.get("name", "")))
    return PlainTextResponse("File tidak ditemukan.", status_code=404)


# =============================================================
# ROUTE UTAMA — GET/POST / (ENDPOINT = location.pathname)
# =============================================================


PAGE_HTML = None    # landing page (index.html)
TOOLS_HTML = None   # halaman tool analisis (tools.html, dulu index.php)


def chat_llm(messages):
    """Balas chat memakai LLM cloud (OpenAI/Gemini/Azure) via llm_client.
    Menyuntik konteks pengetahuan analis (glosarium/disambiguasi/peta intent)
    yang relevan dengan pesan terakhir user."""
    system = (
        "Kamu asisten AI berbahasa Indonesia yang membantu, ramah, ringkas, dan "
        "akurat. Gunakan format Markdown bila relevan."
    )
    try:
        last_user = ""
        for m in reversed(messages or []):
            if (m.get("role") or "").lower() == "user":
                last_user = m.get("content") or ""
                break
        system += kctx.system_suffix(last_user)
    except Exception:
        pass
    # Masking PII pada seluruh pesan + system sebelum ke LLM cloud (Fase A).
    safe_messages = [dict(m, content=pii_mask.mask_text(m.get("content")))
                     for m in (messages or [])]
    return llm_client.chat(safe_messages, system=pii_mask.mask_text(system),
                           max_new_tokens=1024, temperature=0.2)


async def landing(request: Request):
    """Landing page: chat + akses tools."""
    return render_page(request, "index.html", "")


async def tools(request: Request):
    """Tool analisis Dialogflow + Avaya (Step 1..16)."""
    global TOOLS_HTML
    action = request.query_params.get("action", "")
    if action == "":
        return render_page(request, "tools.html", "tools")
    if action == "download":
        return await run_in_threadpool(handle_download, CONFIG, request)
    try:
        ctx = await build_ctx(request)
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)})
    try:
        # dispatch memakai requests (blocking) untuk proxy ke backend; jalankan di
        # threadpool supaya TIDAK membekukan event loop (bikin halaman muter).
        result = await run_in_threadpool(dispatch, action, CONFIG, ctx)
        if not isinstance(result, dict):
            result = {"result": result}
        if "ok" not in result:
            result["ok"] = True
        # Selalu beri tahu frontend dataset (run) kanonik yang dipakai.
        result.setdefault("run", ctx.run)
        return JSONResponse(result)
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)})


async def api_config():
    provider = (os.environ.get("LLM_PROVIDER", "openai") or "openai").strip().lower()
    if provider.startswith("azure"):
        model = os.environ.get("AZURE_OPENAI_DEPLOYMENT", "gpt-4o-mini")
        label = "Azure OpenAI"
    elif provider in ("gemini", "google"):
        model = os.environ.get("GEMINI_MODEL", "gemini-1.5-flash")
        label = "Google Gemini"
    else:
        model = os.environ.get("OPENAI_MODEL", "gpt-4o-mini")
        label = "OpenAI"
    return {"ok": True, "provider": provider, "model": model, "label": label}


async def api_chat(request: Request):
    try:
        body = await request.json()
    except Exception:
        body = {}
    messages = body.get("messages") if isinstance(body, dict) else None
    if not isinstance(messages, list) or not messages:
        return JSONResponse({"ok": False, "error": "messages kosong."})
    messages = messages[-20:]  # batasi konteks agar hemat token
    try:
        reply = await run_in_threadpool(chat_llm, messages)
        return JSONResponse({"ok": True, "reply": reply})
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)})


def register(app):
    app.add_api_route("/", landing, methods=["GET"])
    app.add_api_route("/tools", tools, methods=["GET", "POST"])
    app.add_api_route("/api/config", api_config, methods=["GET"])
    app.add_api_route("/api/chat", api_chat, methods=["POST"])
