# -*- coding: utf-8 -*-
"""
reset_verdict_mkta.py
---------------------
Mengosongkan putusan LLM Step 8 pada sheet 'QA Conf MKTA' supaya baris bisa
DINILAI ULANG (mis. setelah pengetahuan SBERT diperbarui).

KENAPA PERLU:
  Step 8 bersifat incremental -- backend melewati baris yang PUTUSAN-nya sudah
  terisi, dan pipeline (step8_run) memakai ulang artifact Step 8 sebelumnya
  (step8.xlsx). Akibatnya menekan "Lempar ke Qwen" lagi menghasilkan
  "Diproses 0 baris". Skrip ini mengosongkan kolom PUTUSAN / INTENT SEHARUSNYA
  / ALASAN sehingga Step 8 memprosesnya kembali.

PEMAKAIAN:
  # kosongkan HANYA baris di bawah ambang (disarankan, mis. < 0.7):
  python reset_verdict_mkta.py _runs/<RUN_ID>/step8.xlsx --below 0.7

  # kosongkan SEMUA baris yang sudah ada putusannya:
  python reset_verdict_mkta.py _runs/<RUN_ID>/step8.xlsx --all

Catatan:
  - Membuat backup <file>.bak sebelum menimpa.
  - Jalankan pada artifact Step 8 milik run terkait (nama file di server:
    step8.xlsx; nama unduhannya 'hasil_putusan_mkta.xlsx'). Setelah itu jalankan
    ulang Step 8 dari UI (mode 'Otomatis (hasil Step 7)').
  - Alternatif reset penuh Step 8: hapus step8.xlsx lalu jalankan Step 8 lagi
    (pipeline akan jatuh ke hasil Step 7 yang masih kosong putusannya).
"""
import sys
import shutil

from openpyxl import load_workbook

SHEET = "QA Conf MKTA"
CLEAR_COLS = ["PUTUSAN", "INTENT SEHARUSNYA", "ALASAN"]
SCORE_COL = "Skor Pemrosesan Bahasa"


def _col_idx(ws, name):
    for c in range(1, (ws.max_column or 0) + 1):
        v = ws.cell(row=1, column=c).value
        if v is not None and str(v).strip().lower() == name.lower():
            return c
    return -1


def _to_float(v):
    try:
        return float(str(v))
    except Exception:
        return None


def main(argv):
    if len(argv) < 2:
        print("Pemakaian: python reset_verdict_mkta.py <file.xlsx> [--below 0.7 | --all]")
        return 2
    path = argv[1]
    mode_all = "--all" in argv
    below = None
    if "--below" in argv:
        try:
            below = float(argv[argv.index("--below") + 1])
        except Exception:
            print("Nilai --below tidak valid.")
            return 2
    if not mode_all and below is None:
        print("Pilih salah satu: --below <ambang> atau --all.")
        return 2

    wb = load_workbook(path)
    if SHEET not in wb.sheetnames:
        print("Sheet '%s' tidak ditemukan di %s." % (SHEET, path))
        return 1
    ws = wb[SHEET]

    clear_idx = [(_n, _col_idx(ws, _n)) for _n in CLEAR_COLS]
    clear_idx = [(n, i) for (n, i) in clear_idx if i > 0]
    if not any(n == "PUTUSAN" for n, _ in clear_idx):
        print("Kolom PUTUSAN tidak ada; tidak ada yang dikosongkan.")
        return 1
    i_score = _col_idx(ws, SCORE_COL)
    if below is not None and i_score < 0:
        print("Kolom '%s' tidak ada; tidak bisa memfilter --below." % SCORE_COL)
        return 1
    i_put = dict(clear_idx)["PUTUSAN"]

    shutil.copyfile(path, path + ".bak")

    cleared = 0
    scanned = 0
    for row in range(2, (ws.max_row or 1) + 1):
        put = ws.cell(row=row, column=i_put).value
        if put is None or str(put).strip() == "":
            continue  # belum ada putusan; lewati
        if below is not None:
            sc = _to_float(ws.cell(row=row, column=i_score).value)
            if sc is None or sc >= below:
                continue
        scanned += 1
        for _, ci in clear_idx:
            ws.cell(row=row, column=ci, value=None)
        cleared += 1

    wb.save(path)
    scope = ("semua baris berputusan" if mode_all else ("baris < %.2f" % below))
    print("OK. Dikosongkan %d baris (%s). Kolom: %s." % (
        cleared, scope, ", ".join(n for n, _ in clear_idx)))
    print("Backup: %s.bak" % path)
    print("Selanjutnya: jalankan ulang Step 8 (mode Otomatis) dari UI.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
