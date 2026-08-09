# -*- coding: utf-8 -*-
"""audit_tp_routes.py — Menu "Audit Training Phrase".

Fitur berdiri sendiri (BUKAN bagian rail Step 1-16 Analisis Dialogflow).
Tujuan: higiene data intent. Menyandingkan training phrase ANTAR intent memakai
SBERT (model sama seperti pipeline) untuk menemukan:
  1) Konflik antar-intent  : frasa mirip yang berada di intent BERBEDA
                             (penyebab utama match rate turun / tumpang tindih).
  2) Duplikat dalam intent : frasa nyaris kembar di intent SAMA (seed mubazir).

Input = ZIP "Database Intent Dialogflow" (output Step 3/13) ATAU langsung file
xlsx Training Phrase berkolom ["ID", "Training Phrase"].

Daftarkan dengan:  import audit_tp_routes; audit_tp_routes.register(app)

RBAC: path /audit-tp & /api/audit-tp/* jatuh ke area "dialogflow" (default),
aksi "read" — sama seperti tool Dialogflow lain, jadi tak perlu ubah izin.
"""
import io
import os
import base64

from fastapi import Request
from fastapi.responses import JSONResponse
from starlette.concurrency import run_in_threadpool

from app_core import render_page

try:
    import numpy as np
except Exception:
    np = None

import openpyxl


def find_latest_step3_zip():
    from app_core import CONFIG
    import pipeline_routes
    runs_dir = CONFIG.get("runs_dir", "")
    if not runs_dir or not os.path.isdir(runs_dir):
        return None
    latest_run = None
    latest_time = ""
    latest_zip = None
    try:
        for run_id in os.listdir(runs_dir):
            state = pipeline_routes.load_state(CONFIG, run_id)
            if not state:
                continue
            step3 = state.get("steps", {}).get("3")
            if step3 and step3.get("file") and step3.get("status") == "Selesai":
                at = step3.get("at", "")
                if at > latest_time:
                    latest_time = at
                    latest_run = run_id
                    latest_zip = os.path.join(runs_dir, run_id, step3.get("file"))
    except Exception:
        pass
    return latest_zip

# ---------------------------------------------------------------------------
# Ambang default (bisa diubah dari form)
# ---------------------------------------------------------------------------
DEF_MIN_CROSS = float(os.environ.get("AUDIT_TP_MIN_CROSS", "0.85"))
DEF_MIN_DUP = float(os.environ.get("AUDIT_TP_MIN_DUP", "0.97"))
MAX_PAIRS = int(os.environ.get("AUDIT_TP_MAX_PAIRS", "2000"))
MAX_PHRASES = int(os.environ.get("AUDIT_TP_MAX_PHRASES", "8000"))
CHUNK = 256


def _default_encode(texts):
    """Pakai encoder SBERT yang sudah ada (knowledge_semantic). Return NxD
    ternormalisasi (float32) atau None bila model tak tersedia."""
    import knowledge_semantic as ksem
    if not ksem.is_available():
        return None
    return ksem._encode(list(texts))


def _read_training_rows(xlsx_bytes):
    """Baca xlsx Training Phrase -> list[(intent, phrase)]. Cari kolom
    'ID'/'Intent' & 'Training Phrase' dari baris header."""
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    header = None
    for r in rows_iter:
        if r and any(str(c or "").strip() for c in r):
            header = [str(c or "").strip() for c in r]
            break
    if not header:
        raise Exception("Sheet Training Phrase kosong.")
    low = [h.lower() for h in header]

    def _find(cands):
        for i, h in enumerate(low):
            if h in cands:
                return i
        return -1

    ci = _find({"id", "intent", "intent name", "nama intent"})
    cp = _find({"training phrase", "training_phrase", "phrase", "frasa", "training phrases"})
    if ci < 0 or cp < 0:
        raise Exception("Header wajib memuat kolom 'ID' (intent) dan 'Training Phrase'. "
                        "Ditemukan: " + ", ".join(header))
    def _cell(row, idx):
        v = row[idx] if idx < len(row) else None
        return "" if v is None else str(v).strip()

    out = []
    for r in rows_iter:
        if not r:
            continue
        intent = _cell(r, ci)
        phrase = _cell(r, cp)
        if intent and phrase:
            out.append((intent, phrase))
    wb.close()
    return out


def audit_phrases(rows, min_cross=DEF_MIN_CROSS, min_dup=DEF_MIN_DUP,
                  top=MAX_PAIRS, chunk=CHUNK, encode_fn=None):
    """Inti algoritma. rows = list[(intent, phrase)]. encode_fn injectable untuk
    pengujian (default: SBERT). Semua embedding diasumsikan ter-normalisasi
    sehingga cosine = dot product."""
    if np is None:
        raise Exception("numpy tidak tersedia di server.")
    # Dedup pasangan (intent, phrase) identik.
    seen = set()
    uniq = []
    for intent, phrase in rows:
        key = (intent, phrase)
        if key in seen:
            continue
        seen.add(key)
        uniq.append((intent, phrase))
    if len(uniq) > MAX_PHRASES:
        raise Exception("Terlalu banyak training phrase (%d > %d). Pecah per kategori dulu."
                        % (len(uniq), MAX_PHRASES))
    n = len(uniq)
    intents = [x[0] for x in uniq]
    phrases = [x[1] for x in uniq]
    if n < 2:
        return {"ok": True, "n_intents": len(set(intents)), "n_phrases": n,
                "min_cross": min_cross, "min_dup": min_dup,
                "cross_pairs": [], "intent_conflicts": [], "dupes": [],
                "note": "Terlalu sedikit frasa untuk diaudit."}

    enc = encode_fn or _default_encode
    emb = enc(phrases)
    if emb is None:
        raise Exception("Model SBERT tidak aktif di server (torch / sentence-transformers). "
                        "Aktifkan KNOWLEDGE_SEMANTIC & pasang dependensi GPU/CPU.")
    emb = np.asarray(emb, dtype="float32")

    cross_seen = set()
    cross_pairs = []
    dupes = []
    intent_arr = np.array(intents, dtype=object)

    for a in range(0, n, chunk):
        b = min(a + chunk, n)
        sims = emb[a:b] @ emb.T                     # (chunk x n)
        for local_i in range(b - a):
            i = a + local_i
            srow = sims[local_i]
            srow[i] = -1.0                          # buang self
            same = (intent_arr == intents[i])
            # --- konflik antar-intent (intent berbeda, >= min_cross) ---
            cross_mask = (~same) & (srow >= min_cross)
            for j in np.nonzero(cross_mask)[0]:
                j = int(j)
                key = (i, j) if i < j else (j, i)
                if key in cross_seen:
                    continue
                cross_seen.add(key)
                cross_pairs.append((float(srow[j]), i, j))
            # --- duplikat dalam intent (intent sama, >= min_dup) ---
            dup_mask = same & (srow >= min_dup)
            for j in np.nonzero(dup_mask)[0]:
                j = int(j)
                if j <= i:
                    continue
                dupes.append((float(srow[j]), i, j))

    cross_pairs.sort(key=lambda t: t[0], reverse=True)
    dupes.sort(key=lambda t: t[0], reverse=True)
    truncated = len(cross_pairs) > top
    cross_pairs = cross_pairs[:top]

    cross_out = [{"intent_a": intents[i], "phrase_a": phrases[i],
                  "intent_b": intents[j], "phrase_b": phrases[j],
                  "score": round(s, 4)} for s, i, j in cross_pairs]
    dup_out = [{"intent": intents[i], "phrase_a": phrases[i],
                "phrase_b": phrases[j], "score": round(s, 4)} for s, i, j in dupes[:top]]

    # Agregasi per pasangan intent.
    agg = {}
    for s, i, j in cross_pairs:
        ia, ib = intents[i], intents[j]
        pk = (ia, ib) if ia <= ib else (ib, ia)
        cur = agg.get(pk)
        if cur is None:
            agg[pk] = [1, s]
        else:
            cur[0] += 1
            if s > cur[1]:
                cur[1] = s
    intent_conflicts = [{"intent_a": k[0], "intent_b": k[1],
                         "count": v[0], "max_score": round(v[1], 4)}
                        for k, v in agg.items()]
    intent_conflicts.sort(key=lambda d: (d["count"], d["max_score"]), reverse=True)

    return {"ok": True, "n_intents": len(set(intents)), "n_phrases": n,
            "min_cross": min_cross, "min_dup": min_dup, "truncated": truncated,
            "cross_pairs": cross_out, "intent_conflicts": intent_conflicts,
            "dupes": dup_out}


def build_report_xlsx(result):
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Konflik Antar-Intent"
    ws1.append(["Skor", "Intent A", "Frasa A", "Intent B", "Frasa B"])
    for p in result.get("cross_pairs", []):
        ws1.append([p["score"], p["intent_a"], p["phrase_a"], p["intent_b"], p["phrase_b"]])
    ws2 = wb.create_sheet("Ringkasan Pasangan Intent")
    ws2.append(["Intent A", "Intent B", "Jumlah Konflik", "Skor Maks"])
    for c in result.get("intent_conflicts", []):
        ws2.append([c["intent_a"], c["intent_b"], c["count"], c["max_score"]])
    ws3 = wb.create_sheet("Duplikat Dalam Intent")
    ws3.append(["Skor", "Intent", "Frasa A", "Frasa B"])
    for d in result.get("dupes", []):
        ws3.append([d["score"], d["intent"], d["phrase_a"], d["phrase_b"]])
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def run_audit(file_bytes, min_cross=DEF_MIN_CROSS, min_dup=DEF_MIN_DUP):
    """Terima bytes upload (ZIP Step 3/13 atau xlsx Training Phrase), audit,
    dan lampirkan xlsx laporan (base64)."""
    train_bytes = None
    try:
        import pipeline_routes
        train_bytes, _content = pipeline_routes.extract_training_intent(file_bytes)
    except Exception:
        train_bytes = file_bytes           # fallback: dianggap xlsx langsung
    rows = _read_training_rows(train_bytes)
    if not rows:
        raise Exception("Tidak ada baris training phrase yang terbaca.")
    result = audit_phrases(rows, min_cross=min_cross, min_dup=min_dup)
    try:
        xlsx = build_report_xlsx(result)
        result["xlsx_b64"] = base64.b64encode(xlsx).decode("ascii")
        result["xlsx_name"] = "audit_training_phrase.xlsx"
    except Exception:
        pass
    return result


# ---------------------------------------------------------------------------
# Route
# ---------------------------------------------------------------------------
async def audit_tp_page(request: Request):
    return render_page(request, "audit_tp.html", "audit_tp")


def _to_float(v, default):
    try:
        return float(v)
    except Exception:
        return default


async def api_audit_run(request: Request):
    form = await request.form()
    
    use_latest = (form.get("use_latest") == "1")
    data = None
    
    if use_latest:
        zip_path = find_latest_step3_zip()
        if not zip_path or not os.path.isfile(zip_path):
            return JSONResponse({"ok": False, "error": "Tidak ditemukan data Step 3/13 terakhir. Harap unggah file atau jalankan Step 3 di Analisis Dialogflow."})
        with open(zip_path, "rb") as f:
            data = f.read()
    else:
        up = form.get("file")
        if up is None or not hasattr(up, "read"):
            return JSONResponse({"ok": False, "error": "Unggah file ZIP/xlsx training phrase atau centang Gunakan Data Terakhir."})
        data = await up.read()

    min_cross = _to_float(form.get("min_cross"), DEF_MIN_CROSS)
    min_dup = _to_float(form.get("min_dup"), DEF_MIN_DUP)

    def _run():
        return run_audit(data, min_cross=min_cross, min_dup=min_dup)
    try:
        return JSONResponse(await run_in_threadpool(_run))
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)})


def register(app):
    app.add_api_route("/audit-tp", audit_tp_page, methods=["GET"])
    app.add_api_route("/api/audit-tp/run", api_audit_run, methods=["POST"])
