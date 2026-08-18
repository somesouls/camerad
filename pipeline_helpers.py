# -*- coding: utf-8 -*-
"""pipeline_helpers.py — Helper leaf (state/artefak, Google auth, HTTP, XLSX)
yang dipakai pipeline_routes.py.

PENYIMPANAN (diubah pada migrasi DB): state & artefak tidak lagi disimpan sebagai
folder _runs/<run>/state.json + file Excel. Sekarang SUMBER KEBENARAN ada di
database (pipeline_store.py). Folder _runs/<run>/ tetap dipakai HANYA sebagai
CACHE disk agar kode step lama yang membaca file (open(...)) tetap berjalan tanpa
diubah. 'run' kini berperan sebagai kunci dataset (dkey = start__end__lang).

Helper HTTP/Google/XLSX di bawah TIDAK berubah.
"""
import os
import re
import io
import time
import json
import zipfile
import datetime as _dt

import requests

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from app_core import CONFIG, XLSX_MIME, BASE_DIR

import pipeline_store as pstore


# =============================================================
# State & artifact — kini DB-backed (pipeline_store) + cache disk
# =============================================================
def _store():
    return pstore.init_db(pstore.connect())


def _parse_dkey(run):
    """run diharapkan berformat 'start__end__lang'. Bila tidak, kembalikan blank."""
    parts = str(run or "").split("__")
    if len(parts) == 3 and parts[0] and parts[1]:
        return parts[0], parts[1], (parts[2] or "id")
    return "", "", "id"


def _resolve_dataset(conn, run, create=True):
    """Petakan 'run' (string) -> baris dataset. run adalah dkey."""
    if run is None or str(run).strip() == "":
        return None
    ds = pstore.get_dataset_by_key(conn, str(run))
    if ds or not create:
        return ds
    rs, re_, lg = _parse_dkey(run)
    if rs and re_:
        return pstore.get_or_create_dataset(conn, rs, re_, lg)
    # 'run' bukan format dkey standar (mis. Run ID lama): buat dataset dgn dkey=run.
    now = _dt.datetime.now().isoformat()
    conn.execute(
        "INSERT OR IGNORE INTO datasets(dkey,range_start,range_end,lang,label,status,"
        "ngrok_url,meta,created_at,updated_at) VALUES(?,?,?,?,?,'active','','{}',?,?)",
        (str(run), "", "", lg, str(run), now, now),
    )
    conn.commit()
    return pstore.get_dataset_by_key(conn, str(run))


def run_dir(cfg, run, create=True):
    d = os.path.join(cfg["runs_dir"].rstrip("/"), str(run))
    if create and not os.path.isdir(d):
        os.makedirs(d, exist_ok=True)
    return d


def state_path(cfg, run):
    # Dipertahankan untuk kompatibilitas impor; tidak lagi jadi sumber kebenaran.
    return os.path.join(run_dir(cfg, run), "state.json")


def _materialize(cfg, run, dataset, conn):
    """Tulis bytes artefak DB ke cache disk bila file belum ada / beda ukuran."""
    if not dataset:
        return
    d = run_dir(cfg, run)
    for a in pstore.list_artifacts(conn, dataset["id"]):
        fname = "step%s.%s" % (a["step"], a.get("ext") or "bin")
        p = os.path.join(d, fname)
        try:
            if os.path.isfile(p) and os.path.getsize(p) == (a.get("size") or -1):
                continue
        except Exception:
            pass
        b = pstore.get_artifact_bytes(conn, dataset["id"], a["step"])
        if b is not None:
            with open(p, "wb") as f:
                f.write(b)


def load_state(cfg, run):
    conn = _store()
    try:
        ds = _resolve_dataset(conn, run, create=False)
        if not ds:
            return {"run": run, "created": _dt.datetime.now().isoformat(), "steps": {}}
        _materialize(cfg, run, ds, conn)
        return pstore.build_state(conn, ds)
    finally:
        conn.close()


def save_state(cfg, run, state):
    """Kompat: hanya ngrok_url yang perlu dipersist. 'steps' dikelola save_artifact."""
    conn = _store()
    try:
        ds = _resolve_dataset(conn, run, create=True)
        if ds and state is not None and state.get("ngrok_url") is not None:
            pstore.set_ngrok(conn, ds["id"], state.get("ngrok_url") or "")
    finally:
        conn.close()


def set_step(cfg, run, n, data):
    """Kompat: artefak sudah tersimpan via save_artifact; kembalikan state terbaru."""
    return load_state(cfg, run)


def get_state(cfg, run):
    state = load_state(cfg, run)
    return {
        "run": state.get("run", run),
        "ngrok_url": state.get("ngrok_url", ""),
        "steps": state.get("steps", {}),
    }


def reset_run(cfg, run):
    """Reset = SOFT (arsipkan dataset di DB) + bersihkan cache disk. Data tetap aman."""
    archived = False
    conn = _store()
    try:
        ds = _resolve_dataset(conn, run, create=False)
        if ds:
            pstore.archive_dataset(conn, ds["id"])
            archived = True
    finally:
        conn.close()
    d = run_dir(cfg, run, create=False)
    if os.path.isdir(d):
        for name in os.listdir(d):
            try:
                os.remove(os.path.join(d, name))
            except Exception:
                pass
        try:
            os.rmdir(d)
        except Exception:
            pass
    return {"cleared": True, "archived": archived}


def save_ngrok(cfg, run, url):
    conn = _store()
    try:
        ds = _resolve_dataset(conn, run, create=True)
        if ds:
            pstore.set_ngrok(conn, ds["id"], url or "")
    finally:
        conn.close()


def mime_for_ext(ext):
    ext = (ext or "").lower()
    return {
        "json": "application/json; charset=utf-8",
        "xlsx": XLSX_MIME,
        "zip": "application/zip",
    }.get(ext, "application/octet-stream")


def save_artifact(cfg, run, n, ext, data_bytes, download_name, summary):
    if isinstance(data_bytes, str):
        data_bytes = data_bytes.encode("utf-8")
    mime = mime_for_ext(ext)
    # 1) Simpan ke DB (sumber kebenaran, upsert — versi terbaru menang).
    conn = _store()
    try:
        ds = _resolve_dataset(conn, run, create=True)
        pstore.save_artifact(conn, ds["id"], int(n), ext, download_name, mime, data_bytes, summary)
    finally:
        conn.close()
    # 2) Tulis juga ke cache disk agar kode step yang membaca file tetap jalan.
    fname = "step%s.%s" % (n, ext)
    try:
        with open(os.path.join(run_dir(cfg, run), fname), "wb") as f:
            f.write(data_bytes)
    except Exception:
        pass
    return {
        "status": "done",
        "file": fname,
        "name": download_name,
        "ext": ext,
        "mime": mime,
        "size": len(data_bytes),
        "summary": summary,
        "at": _dt.datetime.now().isoformat(),
    }


# ---- Akses artefak berbasis DB (dipakai handler unduh & step yg regenerasi) ----
def artifact_bytes(cfg, run, n):
    """Ambil bytes artefak step n dari DB (sumber kebenaran)."""
    conn = _store()
    try:
        ds = _resolve_dataset(conn, run, create=False)
        if not ds:
            return None
        return pstore.get_artifact_bytes(conn, ds["id"], int(n))
    finally:
        conn.close()


def artifact_meta(cfg, run, n):
    conn = _store()
    try:
        ds = _resolve_dataset(conn, run, create=False)
        if not ds:
            return None
        return pstore.get_artifact_meta(conn, ds["id"], int(n))
    finally:
        conn.close()


def resolve_input_bytes(cfg, ctx, upload_field, allowed_exts):
    up = ctx.file(upload_field)
    if up is not None:
        data, name = up
        ext = os.path.splitext(name)[1].lstrip(".").lower()
        if allowed_exts and ext not in allowed_exts:
            raise Exception("Format file harus: %s. Diterima: %s" % (", ".join(allowed_exts), name))
        return data, name
    from_step = 0
    try:
        from_step = int(ctx.P("from_step", "0") or "0")
    except Exception:
        from_step = 0
    if from_step > 0:
        state = load_state(cfg, ctx.run)
        src = state["steps"].get(str(from_step))
        if not src or not src.get("file"):
            raise Exception("Hasil Step %d belum tersedia." % from_step)
        if allowed_exts and str(src.get("ext", "")).lower() not in allowed_exts:
            raise Exception("Hasil Step %d bukan format yang dibutuhkan (%s)." % (from_step, ", ".join(allowed_exts)))
        b = artifact_bytes(cfg, ctx.run, from_step)
        if b is None:
            p = os.path.join(run_dir(cfg, ctx.run), src["file"])
            if not os.path.isfile(p):
                raise Exception("File hasil Step %d hilang dari server." % from_step)
            with open(p, "rb") as f:
                b = f.read()
        return b, src.get("name", "")
    raise Exception("Tidak ada input. Unggah file atau pilih hasil step sebelumnya.")


def read_upload(ctx, field, exts, label):
    up = ctx.file(field)
    if up is None:
        raise Exception("File %s wajib diunggah." % label)
    data, name = up
    ext = os.path.splitext(name)[1].lstrip(".").lower()
    if exts and ext not in exts:
        raise Exception("File %s harus berformat %s." % (label, "/".join(exts)))
    return data, name


# =============================================================
# Google auth (service-account JWT -> access token)
# =============================================================
_token_cache = {"token": None, "exp": 0.0}


def google_token(cfg, ctx):
    override = (ctx.P("access_token", "") or "").strip()
    if override != "":
        return override
    now = time.time()
    if _token_cache["token"] and _token_cache["exp"] - 60 > now:
        return _token_cache["token"]
    file = cfg["service_account_file"]
    if not os.path.isfile(file):
        fallback = os.path.join(BASE_DIR, "service-account.json")
        if os.path.isfile(fallback):
            file = fallback
        else:
            raise Exception(
                "service-account.json tidak ditemukan (dicek: '%s' dan '%s') dan "
                "Access Token kosong. Perbaiki PIPELINE_SA_FILE di .env, atau "
                "letakkan service-account.json di folder yang sama dengan web_app.py, "
                "atau tempel Access Token di form." % (cfg["service_account_file"], fallback)
            )
    try:
        from google.oauth2 import service_account
        from google.auth.transport.requests import Request as GRequest
    except Exception:
        raise Exception("Library google-auth belum terpasang. Jalankan: pip install google-auth")
    try:
        creds = service_account.Credentials.from_service_account_file(
            file, scopes=[cfg["google_scope"]]
        )
        creds.refresh(GRequest())
    except Exception as e:
        raise Exception("Gagal meminta token Google: %s" % e)
    if not creds.token:
        raise Exception("Token Google gagal.")
    _token_cache["token"] = creds.token
    try:
        _token_cache["exp"] = creds.expiry.replace(tzinfo=_dt.timezone.utc).timestamp()
    except Exception:
        _token_cache["exp"] = now + 3000
    return creds.token


# =============================================================
# HTTP helper Google API
# =============================================================
def http_post_json(url, body, token, timeout=120):
    try:
        r = requests.post(
            url, json=body,
            headers={"Authorization": "Bearer " + token, "Content-Type": "application/json"},
            timeout=timeout,
        )
    except Exception as e:
        return (0, None, str(e))
    try:
        j = r.json()
    except Exception:
        j = None
    return (r.status_code, j, r.text)


def http_get_json(url, query, token, timeout=120):
    try:
        r = requests.get(
            url, params=query,
            headers={"Authorization": "Bearer " + token},
            timeout=timeout,
        )
    except Exception as e:
        return (0, None, str(e))
    try:
        j = r.json()
    except Exception:
        j = None
    return (r.status_code, j, r.text)


# =============================================================
# Endpoint backend FastAPI (pengganti ngrok / localhost)
# =============================================================
def resolve_api_base(cfg, raw):
    if cfg["force_local_api"]:
        return cfg["local_api_base"].rstrip("/")
    raw = (raw or "").strip()
    if raw == "":
        return cfg["local_api_base"].rstrip("/")
    if not re.match(r"^https?://", raw, re.I):
        raise Exception("URL server harus diawali http:// atau https://")
    return raw.rstrip("/")


def api_endpoint(cfg, raw, suffix):
    base = resolve_api_base(cfg, raw)
    if base.endswith(suffix):
        return base
    return base + suffix


def _api_headers(cfg, extra=None):
    h = {"X-API-Key": cfg["qwen_api_key"], "ngrok-skip-browser-warning": "true"}
    if extra:
        h.update(extra)
    return h


def curl_multipart(cfg, endpoint, files, fields=None):
    """POST multipart, kembalikan (bytes, headers). Wajib balasan file (PK)."""
    multipart = {}
    for field, info in files.items():
        b, name = info[0], info[1]
        multipart[field] = (name, b, XLSX_MIME)
    r = requests.post(endpoint, headers=_api_headers(cfg), files=multipart, data=(fields or {}), timeout=3600)
    hdrs = {k.lower(): v for k, v in r.headers.items()}
    if r.status_code < 200 or r.status_code >= 300:
        raise Exception("Server error (HTTP %d): %s" % (r.status_code, r.text[:300]))
    ctype = hdrs.get("content-type", "")
    if "json" in ctype.lower():
        raise Exception("Server membalas JSON, bukan file XLSX: %s" % r.text[:300])
    content = r.content
    if content[:2] != b"PK":
        peek = re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", content.decode("utf-8", "replace"))).strip()
        raise Exception("Server tidak mengembalikan file XLSX yang valid (mungkin halaman error/interstitial). Cuplikan: " + peek[:300])
    return content, hdrs


def curl_multipart_raw(cfg, endpoint, files, fields=None):
    """POST multipart, kembalikan bytes mentah (tak paksa PK/JSON)."""
    multipart = {}
    for field, info in files.items():
        b, name = info[0], info[1]
        mime = info[2] if len(info) > 2 else "application/octet-stream"
        multipart[field] = (name, b, mime)
    r = requests.post(endpoint, headers=_api_headers(cfg), files=multipart, data=(fields or {}), timeout=3600)
    if r.status_code < 200 or r.status_code >= 300:
        raise Exception("Server error (HTTP %d): %s" % (r.status_code, r.text[:1500]))
    return r.content


def curl_json_raw(cfg, endpoint, json_str):
    r = requests.post(
        endpoint, headers=_api_headers(cfg, {"Content-Type": "application/json"}),
        data=json_str.encode("utf-8"), timeout=3600,
    )
    if r.status_code < 200 or r.status_code >= 300:
        raise Exception("Server error (HTTP %d): %s" % (r.status_code, r.text[:1500]))
    return r.content


def curl_get_raw(cfg, endpoint):
    r = requests.get(endpoint, headers=_api_headers(cfg), timeout=60)
    if r.status_code < 200 or r.status_code >= 300:
        raise Exception("Server error (HTTP %d): %s" % (r.status_code, r.text[:1500]))
    return r.content


def curl_post_json(cfg, endpoint, files, fields=None):
    multipart = {}
    for field, info in files.items():
        b, name = info[0], info[1]
        mime = info[2] if len(info) > 2 else "application/octet-stream"
        multipart[field] = (name, b, mime)
    r = requests.post(endpoint, headers=_api_headers(cfg), files=multipart, data=(fields or {}), timeout=3600)
    if r.status_code < 200 or r.status_code >= 300:
        raise Exception("Server error (HTTP %d): %s" % (r.status_code, r.text[:300]))
    try:
        return r.json()
    except Exception:
        peek = re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", r.text)).strip()
        raise Exception("Server tidak mengembalikan JSON valid (mungkin halaman error/interstitial). Cuplikan: " + peek[:300])


# =============================================================
# XLSX helper (pakai openpyxl — fungsional identik dgn versi PHP)
# =============================================================
def xlsx_build(sheets):
    """sheets: list of {name, rows(aoa), widths?, wrapCols?(0-based idx)}."""
    wb = Workbook()
    wb.remove(wb.active)
    for sh in sheets:
        ws = wb.create_sheet(title=str(sh["name"])[:31])
        wrap_cols = set(sh.get("wrapCols") or [])
        for r_i, row in enumerate(sh["rows"], start=1):
            for c_i, val in enumerate(row, start=1):
                cell_val = None if (val == "" or val is None) else val
                cell = ws.cell(row=r_i, column=c_i, value=cell_val)
                if r_i == 1:
                    cell.font = Font(bold=True)
                elif (c_i - 1) in wrap_cols:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.freeze_panes = "A2"
        for i, w in enumerate(sh.get("widths") or [], start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def zip_build(files):
    """files: list of {name, data(bytes)}."""
    bio = io.BytesIO()
    with zipfile.ZipFile(bio, "w", zipfile.ZIP_DEFLATED) as z:
        for f in files:
            data = f["data"]
            if isinstance(data, str):
                data = data.encode("utf-8")
            z.writestr(f["name"], data)
    return bio.getvalue()


def _wb_from_bytes(b):
    return load_workbook(io.BytesIO(b), data_only=True)


def sheet_headers(ws):
    H = {}
    for c in range(1, (ws.max_column or 0) + 1):
        v = ws.cell(row=1, column=c).value
        if v is None:
            continue
        name = str(v).strip()
        if name != "":
            H[name] = c
    return H


def read_sheet(ws):
    """Kembalikan {headers:{name:col}, rows:{rownum:{col:value}}, maxRow}."""
    rows = {}
    maxrow = ws.max_row or 0
    maxcol = ws.max_column or 0
    for r in range(1, maxrow + 1):
        cells = {}
        for c in range(1, maxcol + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            cells[c] = v
        if cells:
            rows[r] = cells
    headers = {}
    if 1 in rows:
        for c, v in rows[1].items():
            nm = str(v).strip()
            if nm != "":
                headers[nm] = c
    return {"headers": headers, "rows": rows, "maxRow": maxrow}


def _sv(cells, col):
    """String value dari sel (mirip perilaku PHP yang selalu string)."""
    if not col or col not in cells:
        return ""
    v = cells[col]
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def _is_numeric(v):
    if v is None or v == "":
        return False
    try:
        float(str(v))
        return True
    except Exception:
        return False


def _find_header(headers, cands):
    for c in cands:
        for name, idx in headers.items():
            if str(name).strip().lower() == c.lower():
                return idx
    return None


def xlsx_upsert_sheet(src_bytes, sheet_name, aoa):
    """Buat/timpa 1 worksheet di workbook (mempertahankan sheet lain)."""
    wb = load_workbook(io.BytesIO(src_bytes))
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=str(sheet_name)[:31])
    for r_i, row in enumerate(aoa, start=1):
        for c_i, val in enumerate(row, start=1):
            cell_val = None if (val == "" or val is None) else val
            ws.cell(row=r_i, column=c_i, value=cell_val)
    ws.freeze_panes = "A2"
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
